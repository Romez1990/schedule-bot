from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from renderer import Schedule, Group, GroupSchedule, WeekDay, DaySchedule, \
    Entry


class Parser:
    def parse(self, file_path: Path) -> Schedule:
        sheets = self._load_file(file_path)
        schedule = Schedule()
        for sheet in sheets:
            week_days = self._parse_week_days(sheet)
            self._parse_entries(schedule, sheet, week_days)
        return schedule

    def _load_file(self, file_path: Path) -> List[Worksheet]:
        work_book: Workbook = load_workbook(file_path)
        sheet_names = work_book.sheetnames
        sheets = [work_book[sheet_name] for sheet_name in sheet_names]
        return sheets

    class WeekDayRow:
        def __init__(self, start: int):
            self.start = start
            self.end = -1

    def _parse_week_days(self, sheet: Worksheet) -> Dict[WeekDay, WeekDayRow]:
        week_days: Dict[WeekDay, Parser.WeekDayRow] = {}
        previous_row: Optional[Parser.WeekDayRow] = None
        for row_number in range(1, sheet.max_row + 1):
            week_day_cell = str(sheet.cell(row_number, 1).value)
            try:
                week_day = WeekDay.from_text(week_day_cell)
            except KeyError:
                continue
            week_day_row = self.WeekDayRow(row_number)
            week_days[week_day] = week_day_row

            if previous_row is not None:
                previous_row.end = week_day_row.start - 1

            previous_row = week_day_row
        if previous_row is not None:
            previous_row.end = sheet.max_row
        return week_days

    def _parse_entries(self, schedule: Schedule, sheet: Worksheet,
                       week_days: Dict[WeekDay, WeekDayRow]) -> None:
        first_week_day = list(week_days.values())[0]
        column_start = 5
        group_name = sheet.cell(column_start, first_week_day.start - 2).value
        if group_name is None:
            column_start = 6

        for column in range(column_start, sheet.max_column + 1, 2):
            group_name = sheet.cell(first_week_day.start - 2, column).value
            if group_name is None:
                continue
            group_name = str(group_name)

            group_schedule = GroupSchedule()
            group = Group(group_name)

            for week_day, week_day_row in week_days.items():
                day_schedule = DaySchedule()
                for row_number in range(week_day_row.start,
                                        week_day_row.end, 3):
                    class_number_cell = \
                        sheet.cell(row_number, column).value
                    if class_number_cell is None:
                        if not day_schedule:
                            day_schedule.append(None)
                            continue
                        else:
                            break
                    entry = Entry()
                    entry.subject = str(
                        sheet.cell(row_number, column).value)
                    entry.kind = str(
                        sheet.cell(row_number + 1, column).value)
                    entry.teacher = str(
                        sheet.cell(row_number + 2, column).value)
                    entry.class_room = str(
                        sheet.cell(row_number, column + 1).value)
                    day_schedule.append(entry)
                if day_schedule:
                    group_schedule[week_day] = day_schedule

            schedule[group] = group_schedule
